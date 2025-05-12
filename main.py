import os
import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import sys
import copy

def main():
    # 1. Obtener credenciales y ruta del archivo
    db_name = os.environ.get('DB_NAME')
    db_user = os.environ.get('DB_USER')
    db_password = os.environ.get('DB_PASSWORD')
    db_host = os.environ.get('DB_HOST')
    db_port = os.environ.get('DB_PORT')
    file_path = os.environ.get('EXCEL_FILE_PATH')
    
    db_params = {
        'dbname': db_name,
        'user': db_user,
        'password': db_password,
        'host': db_host,
        'port': db_port
    }
    
    # 2. Definir la nueva consulta SQL con fechas dinámicas
    fecha_inicio_str = '2025-01-01'
    fecha_fin = datetime.now().date()
    fecha_fin_str = fecha_fin.strftime('%Y-%m-%d')
    
    query = f"""
    SELECT 
    rc.name as "Compañía",
    a.internal_number as "Número",
    rp.vat as "CIF",
    p.name AS "Cliente/Proveedor",
	CASE 
    WHEN a.type = 'out_invoice' THEN 'Cliente'
    WHEN a.type = 'in_invoice' THEN 'Proveedor'
    WHEN a.type = 'out_refund' THEN 'Rectificativa Cliente'
    WHEN a.type = 'in_refund' THEN 'Rectificativa Proveedor'
    ELSE 'Otro'
    END as "Tipo",
 
    EXTRACT(DAY FROM a.date_invoice) AS "Fecha Dia",
	EXTRACT(MONTH FROM a.date_invoice) AS "Fecha Mes",
	EXTRACT(YEAR FROM a.date_invoice) AS "Fecha Año",
	TO_CHAR(a.date_invoice, 'DD/MM/YYYY') AS "Fecha",
 
    CASE 
        WHEN afp.name = 'Recargo de Equivalencia' THEN 'Recargo de Equivalencia'
        WHEN afp.name = 'Régimen Extracomunitario' THEN 'Régimen Extracomunitario'
        WHEN afp.name IN ('REGIMEN INTRACOMUNITARIO', 'Régimen Intracomunitario') THEN 'Régimen Intracomunitario'
        WHEN afp.name = 'REGIMEN NACIONAL' THEN 'Régimen Nacional'
        ELSE afp.name
    END as "Régimen fiscal",
 
    a.amount_untaxed_0 as "Base 0", 
    a.amount_taxed_0 as "Importe 0", 
    a.amount_total_0 as "Total 0",
 
    a.amount_untaxed_4 as "Base 4", 
    a.amount_taxed_4 as "Importe 4", 
    a.amount_total_4 as "Total 4",
 
    a.amount_untaxed_8 as "Base 8", 
    a.amount_taxed_8 as "Importe 8", 
    a.amount_total_8 as "Total 8",
 
    a.amount_untaxed_10 as "Base 10", 
    a.amount_taxed_10 as "Importe 10", 
    a.amount_total_10 as "Total 10",
 
    a.amount_untaxed_18 as "Base 18", 
    a.amount_taxed_18 as "Importe 18", 
    a.amount_total_18 as "Total 18",
 
    a.amount_untaxed_21 as "Base 21", 
    a.amount_taxed_21 as "Importe 21", 
    a.amount_total_21 as "Total 21",
 
    a.recargo_equivalencia_05 as "Recargo 0,5", 
    a.recargo_equivalencia_1 as "Recargo 1", 
    a.recargo_equivalencia_14 as "Recargo 1.4", 
    a.recargo_equivalencia_4 as "Recargo 4",
    a.recargo_equivalencia_52 as "Recargo 5.2",
 
    a.amount_total as "Total",
 
  
	(CASE 							    -- Total Bases (suma de todas las bases) con signo negativo en 'Rectificativa Cliente'
        WHEN a.type = 'out_refund' THEN -- Si es rectificativa cliente
            - (
                COALESCE(a.amount_untaxed_0, 0) +
                COALESCE(a.amount_untaxed_4, 0) +
                COALESCE(a.amount_untaxed_8, 0) +
                COALESCE(a.amount_untaxed_10, 0) +
                COALESCE(a.amount_untaxed_18, 0) +
                COALESCE(a.amount_untaxed_21, 0)
            )
        ELSE -- Para otros tipos de factura, el valor permanece positivo
            (
                COALESCE(a.amount_untaxed_0, 0) +
                COALESCE(a.amount_untaxed_4, 0) +
                COALESCE(a.amount_untaxed_8, 0) +
                COALESCE(a.amount_untaxed_10, 0) +
                COALESCE(a.amount_untaxed_18, 0) +
                COALESCE(a.amount_untaxed_21, 0)
            )
    	END) as "Total BASES",
    CASE   
        WHEN a.obsolescencia = TRUE THEN 'SI' 
        ELSE 'NO' 
    END as "BULK",  -- Columna Obsolescencia (Bulk)
	'S-' || rp.id AS "ID Cliente",
	rpa.city AS "Ciudad"
 
FROM 
    account_invoice as a 
INNER JOIN 
    res_partner as p ON p.id = a.partner_id 
INNER JOIN 
    account_fiscal_position afp ON afp.id = a.fiscal_position 
INNER JOIN 
    res_company rc ON rc.id = a.company_id
INNER JOIN 
    res_partner rp ON rp.id = a.partner_id
INNER JOIN 
	res_partner_address rpa ON rpa.id = a.address_invoice_id
 
WHERE 
    a.state IN ('open','paid') 
    AND a.date_invoice BETWEEN '{fecha_inicio_str}' AND '{fecha_fin_str}'
    AND a.type IN ('out_invoice', 'out_refund') -- Solo facturas y rectificativas de clientes
	AND a.internal_number NOT LIKE 'RA%' -- Omitir RAPPEL
 
ORDER BY 
    a.internal_number;
    """
    
    # 3. Ejecutar la consulta
    try:
        with psycopg2.connect(**db_params) as conn:
            with conn.cursor() as cur:
                cur.execute(query)
                resultados = cur.fetchall()
                headers = [desc[0] for desc in cur.description]
    except Exception as e:
        print(f"Error al conectar o ejecutar la consulta: {e}")
        sys.exit(1)
    
    if not resultados:
        print("No se obtuvieron resultados de la consulta.")
        return
    else:
        print(f"Se obtuvieron {len(resultados)} filas de la consulta.")
    
    # 4. Cargar archivo y buscar tabla existente
    try:
        book = load_workbook(file_path)
        sheet = book.active
        table = sheet.tables.get("Portes")
        if not table:
            print("No se encontró la tabla 'Portes'.")
            return
    except FileNotFoundError:
        print(f"No se encontró el archivo '{file_path}'.")
        return

    # 5. Determinar límites de la tabla existente
    start_cell, end_cell = table.ref.split(':')
    start_col_letter = ''.join(filter(str.isalpha, start_cell))
    end_col_letter = ''.join(filter(str.isalpha, end_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    end_row = int(''.join(filter(str.isdigit, end_cell)))
    start_col = column_index_from_string(start_col_letter)
    end_col = column_index_from_string(end_col_letter)

    # 6. Obtener códigos existentes para evitar duplicados (columna 3 = VAT/CIF)
    existing_invoice_codes = {
        sheet.cell(row=i, column=2).value
        for i in range(start_row + 1, end_row + 1)
        if sheet.cell(row=i, column=3).value is not None
    }

    # 7. Insertar filas justo debajo del final de la tabla
    new_data = []
    for row in resultados:
        if row[2] not in existing_invoice_codes:
            new_data.append(row)
    
    if not new_data:
        print("No hay datos nuevos que añadir.")
        return
    
    insert_row = end_row + 1
    for row in new_data:
        for col_index in range(len(row)):
            cell = sheet.cell(row=insert_row, column=start_col + col_index, value=row[col_index])
            # Aplicar estilo desde la última fila válida
            source_cell = sheet.cell(row=end_row, column=start_col + col_index)
            cell.font = copy.copy(source_cell.font)
            cell.fill = copy.copy(source_cell.fill)
            cell.border = copy.copy(source_cell.border)
            cell.alignment = copy.copy(source_cell.alignment)
        insert_row += 1

    # 8. Actualizar rango de la tabla
    new_end_row = end_row + len(new_data)
    new_ref = f"{start_col_letter}{start_row}:{end_col_letter}{new_end_row}"
    table.ref = new_ref
    print(f"Tabla 'Portes' actualizada a rango: {new_ref}")

    # 9. Guardar archivo
    book.save(file_path)
    print(f"Archivo guardado en '{file_path}'.")

if __name__ == '__main__':
    main()
